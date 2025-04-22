import requests
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import calendar
import os
import zipfile
import pandas as pd
import logging
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

logging.basicConfig(
    filename='eiopa_rates.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def get_download_link(url):
    """Get the download link for the previous month's EIOPA rates ZIP file"""
    try:
        # Get current date
        today = datetime.now()
        
        # Get previous month (if current month is January, goes to December of previous year)
        if today.month == 1:
            target_year = today.year - 1
            target_month = 12
        else:
            target_year = today.year
            target_month = today.month - 1
        
        # Get last day of the target month
        last_day = calendar.monthrange(target_year, target_month)[1]
        
        # Format the expected filename (YYYYMMDD)
        filename = f"EIOPA_RFR_{target_year}{target_month:02d}{last_day}"
        
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Look for the previous month's download link
        for link in soup.find_all('a', href=True):
            href = link.get('href')
            if ('download' in href.lower() and 
                f'_en?filename={filename}.zip' in href):
                print(f"Found {calendar.month_name[target_month]} {target_year} link: {href}")
                # Ensure the URL is absolute
                return f"https://www.eiopa.europa.eu{href}" if not href.startswith('http') else href
        
        print(f"No download link found for {calendar.month_name[target_month]} {target_year}")
        return None
        
    except Exception as e:
        print(f"Error retrieving page: {e}")
        return None

def download_and_extract_excel(url, download_dir):
    """Download ZIP file and extract the 3rd Excel file"""
    if not os.path.exists(download_dir):
        os.makedirs(download_dir)
    
    temp_zip_path = os.path.join(download_dir, 'temp.zip')
    
    try:
        # Download and save ZIP file
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        with open(temp_zip_path, 'wb') as f:
            f.write(response.content)
        
        # Extract the 3rd Excel file
        with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
            excel_files = sorted([f for f in zip_ref.namelist() if f.endswith('.xlsx')])
            
            if len(excel_files) >= 3:
                third_excel = excel_files[2]
                zip_ref.extract(third_excel, download_dir)
                print(f"Extracted {third_excel}")
                return os.path.join(download_dir, third_excel)
            else:
                print("ZIP file does not contain enough Excel files")
                return None
                
    except Exception as e:
        print(f"Error processing file: {e}")
        return None
    finally:
        # Clean up temp zip file
        if os.path.exists(temp_zip_path):
            try:
                os.unlink(temp_zip_path)
            except Exception:
                pass

def create_and_save_plot(df, excel_path):
    """Create a line plot of Euro rates and save it to both Excel and as PNG"""
    try:
        # Create the plot with smaller figure size
        plt.figure(figsize=(8, 4))  # Reduced from (12, 6)
        
        # Plot lines without markers
        plt.plot(df['Maturity'], df['Euro_Rate_No_VA'], label='No VA', linewidth=1)
        plt.plot(df['Maturity'], df['Euro_Rate_With_VA'], label='With VA', linewidth=1)
        
        # Customize the plot
        plt.title('EIOPA Euro Rates by Maturity', fontsize=10)
        plt.xlabel('Maturity', fontsize=9)
        plt.ylabel('Rate (%)', fontsize=9)
        plt.grid(True, linestyle='--', alpha=0.7)
        plt.legend(fontsize=8)
        plt.tick_params(labelsize=8)
        
        # Make plot more compact
        plt.tight_layout()
        
        # Save as PNG with smaller size
        plot_path = os.path.join(os.path.dirname(excel_path), 'Euro_Rates_Plot.png')
        plt.savefig(plot_path, dpi=200, bbox_inches='tight')
        
        # Add plot to Excel with specific size
        workbook = load_workbook(excel_path)
        worksheet = workbook.create_sheet('Plot')
        
        # Add the image to the worksheet with size control
        img = Image(plot_path)
        img.width = 400  # Adjust width in pixels
        img.height = 200  # Adjust height in pixels
        worksheet.add_image(img, 'A1')
        
        # Save Excel with plot
        workbook.save(excel_path)
        
        logging.info(f"Created plot and saved to {plot_path} and Excel file")
        return True
        
    except Exception as e:
        logging.error(f"Error creating plot: {e}")
        return False

def extract_euro_rates(excel_file):
    """Extract percentage rates and create visualizations"""
    try:
        # Read all sheets
        df_no_va = pd.read_excel(excel_file, sheet_name='RFR_spot_no_VA')
        df_with_va = pd.read_excel(excel_file, sheet_name='RFR_spot_with_VA')
        df_no_va_up = pd.read_excel(excel_file, sheet_name='Spot_NO_VA_shock_UP')
        df_no_va_down = pd.read_excel(excel_file, sheet_name='Spot_NO_VA_shock_DOWN')
        df_with_va_up = pd.read_excel(excel_file, sheet_name='Spot_WITH_VA_shock_UP')
        df_with_va_down = pd.read_excel(excel_file, sheet_name='Spot_WITH_VA_shock_DOWN')
        
        # Get column C starting from the 10th row (index 9) for all sheets
        euro_rates_no_va = df_no_va.iloc[9:, 2].reset_index(drop=True)
        euro_rates_with_va = df_with_va.iloc[9:, 2].reset_index(drop=True)
        euro_rates_no_va_up = df_no_va_up.iloc[9:, 2].reset_index(drop=True)
        euro_rates_no_va_down = df_no_va_down.iloc[9:, 2].reset_index(drop=True)
        euro_rates_with_va_up = df_with_va_up.iloc[9:, 2].reset_index(drop=True)
        euro_rates_with_va_down = df_with_va_down.iloc[9:, 2].reset_index(drop=True)
        
        # Create output DataFrame
        output_df = pd.DataFrame({
            'Maturity': range(1, len(euro_rates_no_va) + 1),
            'Euro_Rate_No_VA': euro_rates_no_va,
            'Euro_Rate_With_VA': euro_rates_with_va,
            'Euro_Rate_No_VA_Up': euro_rates_no_va_up,
            'Euro_Rate_No_VA_Down': euro_rates_no_va_down,
            'Euro_Rate_With_VA_Up': euro_rates_with_va_up,
            'Euro_Rate_With_VA_Down': euro_rates_with_va_down
        })
        
        # Save to Excel
        output_path = os.path.join(os.path.dirname(excel_file), 'Euro_Rates_Percentages.xlsx')
        output_df.to_excel(output_path, index=False, sheet_name='Percentages')
        
        return output_df
        
    except Exception as e:
        print(f"Error extracting Euro rates: {e}")
        return None

def main():
    """Main execution function"""
    logging.info("Starting EIOPA rates extraction")
    
    url = "https://www.eiopa.europa.eu/tools-and-data/risk-free-interest-rate-term-structures_en"
    zip_file_url = get_download_link(url)

    if zip_file_url:
        logging.info(f"Found download link: {zip_file_url}")
        excel_file = download_and_extract_excel(zip_file_url, 'downloads')
        if excel_file:
            logging.info(f"Extracted Excel file: {excel_file}")
            euro_rates = extract_euro_rates(excel_file)
            if euro_rates is not None:
                logging.info("Successfully extracted EIOPA rates")
                return True
    logging.error("Failed to complete EIOPA rates extraction")
    return False

if __name__ == "__main__":
    main()
